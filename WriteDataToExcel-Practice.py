import openpyxl
file="C:\\Users\\PMLS\\Desktop\\Selenium\\testdata2.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

sheet.cell(1,1).value="Automation Tools"
sheet.cell(1,2).value="Languages"
sheet.cell(1,3).value="Users"

sheet.cell(2,1).value="Selenium"
sheet.cell(2,2).value="Python"
sheet.cell(2,3).value="10000"

sheet.cell(3,1).value="Cypress"
sheet.cell(3,2).value="Javascript"
sheet.cell(3,3).value="6543"

sheet.cell(4,1).value="Playwright"
sheet.cell(4,2).value="Java"
sheet.cell(4,3).value="321"

workbook.save(file)