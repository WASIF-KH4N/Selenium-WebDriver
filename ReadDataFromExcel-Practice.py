import openpyxl


file="C:\\Users\\PMLS\\Desktop\\Selenium\\testdata1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet2"]

row=sheet.max_row # count number of rows in the sheet
column=sheet.max_column # count number of columns in the sheet

for r in range(1,row+1):
    for c in range(1,column+1):
       print(sheet.cell(r,c).value,end="   ")
    print()
